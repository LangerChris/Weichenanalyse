"""Labels, Referenzen und Zieldefinition für die Endlage-Vorhersage.

- Zielcodes: 2724 (Störung Endlage Links), 2723 (Störung Endlage Rechts).
- Frühwarn-Label: ein Umlauf ist positiv, wenn innerhalb der nächsten `horizon`
  Umläufe (je Weiche×Richtung, zeitlich sortiert) eine Endlage-Störung auftritt.
- Statische Weichen-Metadaten (aus *_switches.csv) als Transfer-Features.
- Bestätigte Störungen (Excel) als Gold-Validierung + präzise Vorfalldaten.
"""

from __future__ import annotations

from pathlib import Path

import numpy as np
import pandas as pd

from weichenanalyse.data import DEFAULT_META, load_meta

TARGET_CODES = (2723, 2724)  # 2723 = Endlage R, 2724 = Endlage L
GROUP_KEYS = ["object_id", "position"]


def _switches_path(meta_path: Path) -> Path:
    return meta_path.with_name(meta_path.stem + "_switches.csv")


def load_switches(meta_path: Path | str = DEFAULT_META) -> pd.DataFrame:
    """Statische Weichen-Metadaten (eine Zeile je Weiche)."""
    return pd.read_csv(_switches_path(Path(meta_path)))


def load_error_code_descriptions(
    path: Path | str | None = None,
) -> dict[int, str]:
    """Fehlercode → Beschreibung aus data/reference/error_codes.csv (falls vorhanden)."""
    if path is None:
        path = DEFAULT_META.parent / "reference" / "error_codes.csv"
    path = Path(path)
    if not path.exists():
        return {}
    # In Excel gespeicherte CSVs sind oft cp1252, nicht UTF-8.
    try:
        df = pd.read_csv(path, encoding="utf-8")
    except UnicodeDecodeError:
        df = pd.read_csv(path, encoding="cp1252")
    out = {}
    for _, r in df.iterrows():
        desc = str(r.get("beschreibung") or "").strip()
        if desc and not pd.isna(r.get("code_id")):
            out[int(r["code_id"])] = desc
    return out


def load_confirmed_faults(
    xlsx: Path | str | None = None,
    switches: pd.DataFrame | None = None,
    meta_path: Path | str = DEFAULT_META,
) -> pd.DataFrame:
    """Bestätigte Störungen aus der Excel, gemappt auf object_id (über HAR-Dateiname)."""
    if xlsx is None:
        xlsx = DEFAULT_META.parent / "labels" / "stoerungen.xlsx"
    xlsx = Path(xlsx)
    if not xlsx.exists():
        return pd.DataFrame()
    if switches is None:
        switches = load_switches(meta_path)

    from openpyxl import load_workbook

    wb = load_workbook(xlsx, read_only=True, data_only=True)
    ws = wb["Stoerungen"]
    rows = list(ws.iter_rows(values_only=True))
    # Zeile 0 = Header, Zeile 1 = Hinweise → Daten ab Zeile 2
    records = []
    for r in rows[2:]:
        if not r or not r[0] or "Dateiname" in str(r[0]):
            continue
        records.append({
            "har_key": str(r[0]).replace(".har", "").strip(),
            "weiche": r[1],
            "datum_beginn": r[2],
            "datum_reparatur": r[3],
            "richtung": r[4],
            "fehlerart": r[5],
            "schweregrad": r[7],
            "sicherheit": r[8],
            "notiz": r[10],
        })
    faults = pd.DataFrame(records)
    if faults.empty:
        return faults

    sw = switches.copy()
    sw["har_key"] = sw["har_file"].str.replace(".har", "", regex=False).str.strip()
    return faults.merge(sw[["har_key", "object_id", "label"]], on="har_key", how="left")


def add_target_labels(
    meta: pd.DataFrame,
    target_codes: tuple[int, ...] = TARGET_CODES,
    horizon: int = 0,
) -> pd.DataFrame:
    """Ziel- und Frühwarn-Label ergänzen.

    `is_target`     : dieser Umlauf trägt selbst einen Endlage-Code.
    `is_prefailure` : innerhalb der nächsten `horizon` Umläufe (inkl. diesem,
                      je Weiche×Richtung zeitlich sortiert) folgt eine Endlage-Störung.
    Mit horizon=0 ist `is_prefailure` == `is_target`.
    """
    tc = set(target_codes)
    out = meta.copy()
    out["is_target"] = out["error_ids"].apply(lambda L: bool(tc & set(L)))
    out["is_prefailure"] = False

    for _, idx in out.groupby(GROUP_KEYS).groups.items():
        g = out.loc[idx].sort_values("time")
        tgt = g["is_target"].to_numpy()
        order = g.index.to_numpy()
        pre = np.zeros(len(g), dtype=bool)
        # Position der nächsten Ziel-Events
        target_pos = np.where(tgt)[0]
        if target_pos.size:
            for i in range(len(g)):
                # gibt es ein Ziel-Event in [i, i+horizon]?
                nxt = target_pos[target_pos >= i]
                if nxt.size and nxt[0] <= i + horizon:
                    pre[i] = True
        out.loc[order, "is_prefailure"] = pre
    return out


def unify_amplitude(meta: pd.DataFrame) -> pd.DataFrame:
    """Einheitliches Amplituden-Signal: Strom (A) wo vorhanden, sonst Leistung (W).

    Jede Weiche misst nur eine der beiden Größen. Da Warner per-Weiche-relativ (z zur
    Eigen-Baseline) arbeiten, ist die Einheit egal — `mean_amp`/`peak_amp` vereinheitlichen
    den Zugriff. `signal_unit` hält fest, welche Größe es war.
    """
    out = meta.copy()
    mc, pc = out.get("motor_0_mean_current"), out.get("motor_0_mean_power")
    out["mean_amp"] = mc.where(mc.notna(), pc) if mc is not None else pc
    kc, kp = out.get("motor_0_peak_current"), out.get("motor_0_peak_power")
    out["peak_amp"] = kc.where(kc.notna(), kp) if kc is not None else kp
    out["signal_unit"] = np.where(out.get("motor_0_mean_current").notna(), "A", "W")
    return out


def load_labeled_dataset(
    meta_path: Path | str = DEFAULT_META, horizon: int = 0
):
    """Bequemer Einstieg: Meta + Switch-Metadaten + Ziel-/Frühwarn-Label + Amplituden-Signal."""
    meta = load_meta(Path(meta_path))
    meta = add_target_labels(meta, horizon=horizon)
    meta = unify_amplitude(meta)
    switches = load_switches(meta_path)
    return meta.merge(switches, on="object_id", how="left", suffixes=("", "_sw"))
