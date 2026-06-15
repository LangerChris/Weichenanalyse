"""Erzeugt eine Excel-Vorlage zum Erfassen bekannter Weichen-Störungen (Gold-Labels).

Die ausgefüllte Datei dient später als (partielle) Ground-Truth zur Validierung der
Anomalieerkennung. Wichtig: gemeldete Störungen sind High-Precision-Positive — fehlende
Einträge bedeuten NICHT "gesund" (es können unbekannte Störungen enthalten sein).

Usage:
    python scripts/make_label_template.py            # -> data/labels/stoerungen_template.xlsx
    python scripts/make_label_template.py -o pfad.xlsx
"""

from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

REPO_ROOT = Path(__file__).resolve().parent.parent

# Bekannte Fehlerarten aus DIANA (als Hilfe/Dropdown; "Sonstige/Unbekannt" erlaubt freie Eingabe).
FEHLERARTEN = [
    "Erhöhter Strom/Leistung Umlauf",
    "Stark erhöhter Strom/Leistung Umlauf",
    "Erhöhter Strom/Leistung Verriegelung",
    "Stark erhöhter Strom/Leistung Verriegelung",
    "Erhöhter Strom/Leistung Entriegelung",
    "Stark erhöhter Strom/Leistung Entriegelung",
    "Störung Endlage",
    "Stark verlängerte Umlaufzeit",
    "Abweichende Referenz",
    "Sonstige/Unbekannt",
]
RICHTUNG = ["L", "R", "beide", "unbekannt"]
SCHWEREGRAD = ["Warnung", "Störung", "Ausfall", "unbekannt"]
SICHERHEIT = ["bestätigt", "vermutet"]

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF")
HINT_FONT = Font(italic=True, color="808080")

# Spalten der Störungs-Tabelle: (Überschrift, Breite, Pflicht?, Hinweis)
STOERUNG_COLS = [
    ("HAR_Datei", 22, True, "Dateiname der zugehörigen HAR, z.B. WE265.har"),
    ("Weiche", 18, True, "Klartext-Name, z.B. WE265 / Bf Höchst"),
    ("Datum_Beginn", 16, True, "Wann die Störung begann (JJJJ-MM-TT)"),
    ("Datum_Reparatur", 16, False, "Wann behoben/repariert (falls bekannt)"),
    ("Richtung", 12, False, "L, R, beide oder unbekannt"),
    ("Fehlerart", 30, True, "Kategorie (Dropdown) oder frei"),
    ("Komponente", 22, False, "z.B. Antrieb SAT01, falls bekannt"),
    ("Schweregrad", 14, False, "Warnung / Störung / Ausfall"),
    ("Sicherheit", 14, True, "bestätigt oder vermutet"),
    ("Quelle", 18, False, "Woher die Info stammt (DIANA, Instandhaltung, ...)"),
    ("Notizen", 40, False, "Freitext"),
]

GESUND_COLS = [
    ("HAR_Datei", 22, "Dateiname der zugehörigen HAR"),
    ("Weiche", 18, "Klartext-Name"),
    ("Zeitraum_von", 16, "Bestätigt gesund ab (JJJJ-MM-TT)"),
    ("Zeitraum_bis", 16, "Bestätigt gesund bis (JJJJ-MM-TT)"),
    ("Quelle", 18, "Woher bestätigt"),
    ("Notizen", 40, "Freitext"),
]


def _style_header(ws, headers_widths):
    for i, (name, width, *_rest) in enumerate(headers_widths, start=1):
        col = get_column_letter(i)
        cell = ws.cell(row=1, column=i, value=name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center")
        ws.column_dimensions[col].width = width
    # Hinweiszeile (Zeile 2)
    for i, item in enumerate(headers_widths, start=1):
        hint = item[-1]
        c = ws.cell(row=2, column=i, value=hint)
        c.font = HINT_FONT
    ws.freeze_panes = "A3"


def _add_dropdown(ws, col_letter, options, first=3, last=500):
    dv = DataValidation(
        type="list",
        formula1='"' + ",".join(options) + '"',
        allow_blank=True,
        showDropDown=False,
    )
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}{first}:{col_letter}{last}")


def build_workbook() -> Workbook:
    wb = Workbook()

    # --- Sheet 1: Anleitung ---
    ws = wb.active
    ws.title = "Anleitung"
    lines = [
        ("So füllst du diese Vorlage aus", True),
        ("", False),
        ("1. Tab 'Stoerungen': eine Zeile je bekannter Störung einer Weiche.", False),
        ("   HAR_Datei muss zum Dateinamen in data/Umlaeufe/ passen (objectId wird automatisch zugeordnet).", False),
        ("2. Datumsangaben als JJJJ-MM-TT (z.B. 2025-03-14).", False),
        ("3. Dropdown-Felder (Richtung, Fehlerart, Schweregrad, Sicherheit) nutzen; 'Sonstige/Unbekannt' geht auch.", False),
        ("4. Nur eintragen, was du WEISST. Unbekannte Felder leer lassen.", False),
        ("", False),
        ("WICHTIG: Gemeldete Störungen sind 'sichere Treffer'. Fehlt eine Zeile, heißt das NICHT,", False),
        ("dass die Weiche dort gesund war — es können unbekannte Störungen enthalten sein.", False),
        ("", False),
        ("5. Tab 'Gesund_bestaetigt' (optional): Zeiträume, in denen eine Weiche sicher gesund war.", False),
        ("   Das stärkt die Validierung, ist aber freiwillig.", False),
        ("", False),
        ("Ablage: ausgefüllte Datei nach data/labels/ legen (z.B. data/labels/stoerungen.xlsx).", False),
    ]
    for r, (text, bold) in enumerate(lines, start=1):
        c = ws.cell(row=r, column=1, value=text)
        if bold:
            c.font = Font(bold=True, size=13)
    ws.column_dimensions["A"].width = 110

    # --- Sheet 2: Stoerungen ---
    ws2 = wb.create_sheet("Stoerungen")
    _style_header(ws2, STOERUNG_COLS)
    # Dropdowns: Richtung(E), Fehlerart(F), Schweregrad(H), Sicherheit(I)
    _add_dropdown(ws2, "E", RICHTUNG)
    _add_dropdown(ws2, "F", FEHLERARTEN)
    _add_dropdown(ws2, "H", SCHWEREGRAD)
    _add_dropdown(ws2, "I", SICHERHEIT)

    # --- Sheet 3: Gesund_bestaetigt ---
    ws3 = wb.create_sheet("Gesund_bestaetigt")
    _style_header(ws3, GESUND_COLS)

    # --- Sheet 4: Referenz Fehlerarten ---
    ws4 = wb.create_sheet("Referenz_Fehlerarten")
    ws4.cell(row=1, column=1, value="Bekannte Fehlerarten (aus DIANA)").font = Font(bold=True)
    for r, fa in enumerate(FEHLERARTEN, start=2):
        ws4.cell(row=r, column=1, value=fa)
    ws4.column_dimensions["A"].width = 45

    return wb


def main():
    parser = argparse.ArgumentParser(description="Excel-Vorlage für Störungs-Labels erzeugen")
    parser.add_argument(
        "-o", "--output",
        default=str(REPO_ROOT / "data" / "labels" / "stoerungen_template.xlsx"),
        help="Ausgabepfad der .xlsx",
    )
    args = parser.parse_args()

    out = Path(args.output)
    out.parent.mkdir(parents=True, exist_ok=True)
    build_workbook().save(out)
    print(f"Vorlage geschrieben: {out}")


if __name__ == "__main__":
    main()
