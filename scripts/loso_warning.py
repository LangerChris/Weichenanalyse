"""Leave-one-switch-out-Validierung des Ensemble-Warners.

Ehrliche Generalisierung: Die globalen Schwellen (z, min_consecutive, tail) werden je
Durchlauf auf ALLEN Weichen AUSSER einer gewählt (max. Recall minus Strafe für Fehlalarme)
und dann auf der ausgelassenen Weiche getestet. Die per-Weiche-Baseline ist selbst-
referenziell (nur Eigenhistorie) → kein Leakage über Weichen.

Usage:
    python scripts/loso_warning.py
    python scripts/loso_warning.py --fa-penalty 0.5
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import numpy as np
import pandas as pd
from openpyxl import load_workbook

from weichenanalyse.labels import DEFAULT_META, load_confirmed_faults, load_labeled_dataset, load_switches
from weichenanalyse.shape import SHAPE_COLS, attach_shape_features
from weichenanalyse.warning import EnsembleWarning, alarm_active

FEATURES = ("mean_amp", "peak_amp", "turn_time", *SHAPE_COLS)
Z_GRID = (1.0, 1.5, 2.0, 2.5)
MIN_GRID = (3, 5, 10)
TAIL_GRID = (0, 20, 40)


def healthy_object_ids(switches):
    ws = load_workbook(DEFAULT_META.parent / "labels" / "stoerungen.xlsx", data_only=True)["Gesund_bestaetigt"]
    sw = switches.copy()
    sw["key"] = sw["har_file"].str.replace(".har", "", regex=False).str.strip()
    keys = [str(r[0]).replace(".har", "").strip()
            for r in list(ws.iter_rows(values_only=True))[2:]
            if r[0] and "Dateiname" not in str(r[0])]
    return list(sw[sw.key.isin(keys)].object_id)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--fa-penalty", type=float, default=0.5,
                    help="Strafgewicht für Fehlalarm-Rate bei der Schwellenwahl")
    args = ap.parse_args()

    meta = attach_shape_features(load_labeled_dataset(horizon=0))
    meta["dt"] = pd.to_datetime(meta["time"], unit="ms")
    switches = load_switches()
    healthy = healthy_object_ids(switches)
    faults = load_confirmed_faults()
    faults["fd"] = pd.to_datetime(faults["datum_beginn"])
    faults = faults[~faults["notiz"].astype(str).str.contains("Stein", na=False)]  # abrupt raus
    faults = faults[faults.object_id.notna()].drop_duplicates("object_id")

    # Bewertungseinheiten: (oid, is_fault, fault_date)
    units = [(r.object_id, True, r["fd"]) for _, r in faults.iterrows()]
    units += [(oid, False, None) for oid in healthy if oid not in set(faults.object_id)]

    # Vorhersagen je (z, min) einmal cachen (Baseline ist je Weiche selbst-referenziell).
    cache = {}
    for z in Z_GRID:
        for m in MIN_GRID:
            pred = EnsembleWarning(features=FEATURES, z_thresh=z, min_consecutive=m).fit(meta).predict(meta)
            pred = pred[["object_id", "warn"]].assign(dt=meta["dt"].values, time=meta["time"].values)
            by_oid = {oid: g.sort_values("dt") for oid, g in pred.groupby("object_id")}
            cache[(z, m)] = by_oid

    def outcome(z, m, tail, oid, fd):
        g = cache[(z, m)].get(oid)
        if g is None:
            return False
        if fd is not None:
            g = g[g.dt <= fd]
        return alarm_active(g, tail=tail)

    def score_on(train_units, z, m, tail):
        f = [u for u in train_units if u[1]]
        h = [u for u in train_units if not u[1]]
        rec = np.mean([outcome(z, m, tail, oid, fd) for oid, _, fd in f]) if f else 0.0
        fa = np.mean([outcome(z, m, tail, oid, None) for oid, _, fd in h]) if h else 0.0
        return rec - args.fa_penalty * fa, rec, fa

    # LOSO
    results = []
    chosen = []
    for i, u in enumerate(units):
        train = units[:i] + units[i + 1:]
        best, best_cfg = -1e9, None
        for z in Z_GRID:
            for m in MIN_GRID:
                for tail in TAIL_GRID:
                    s, _, _ = score_on(train, z, m, tail)
                    if s > best:
                        best, best_cfg = s, (z, m, tail)
        z, m, tail = best_cfg
        oid, is_fault, fd = u
        flagged = outcome(z, m, tail, oid, fd)
        results.append((oid, is_fault, flagged))
        chosen.append(best_cfg)

    faults_r = [r for r in results if r[1]]
    healthy_r = [r for r in results if not r[1]]
    recall = np.mean([r[2] for r in faults_r])
    fa_rate = np.mean([r[2] for r in healthy_r])
    print("=== Leave-one-switch-out (ehrliche Generalisierung) ===")
    print(f"Bewertungseinheiten: {len(faults_r)} Störungen (graduell) + {len(healthy_r)} gesund")
    print(f"LOSO-Recall (Störungen rechtzeitig gewarnt): {sum(r[2] for r in faults_r)}/{len(faults_r)} "
          f"= {recall:.0%}")
    print(f"LOSO-Fehlalarm (gesunde Weichen): {sum(r[2] for r in healthy_r)}/{len(healthy_r)} = {fa_rate:.0%}")
    from collections import Counter
    print(f"\nGewählte Konfigurationen (z,min,tail): {dict(Counter(chosen))}")


if __name__ == "__main__":
    main()
