"""Lernendes Transfer-Modell für die Endlage-Frühwarnung, Leave-one-switch-out.

- Positive Umläufe = Fenster (Tage) VOR dem bestätigten Vorfall; gesunde Weichen durchweg
  negativ (auch mit gelegentlichen Codes/Peaks — diese sollen NICHT warnen).
- Modell: Gradient Boosting auf per-Weiche normierten Merkmalen + Persistenz + statischen
  Metadaten (weichenanalyse/model.py).
- LOSO: je ausgelassener Weiche auf allen anderen trainieren. Pro-Umlauf-Wahrscheinlichkeit
  → geglättet (Persistenz!) → Weichen-Score. ROC-AUC trennt Störungs- von gesunden Weichen.
- Fehlalarm zählt nur bei ANHALTEND hohem Score (geglättet), nicht bei Einzel-Blips.

Usage:
    python scripts/loso_model.py --horizon-days 30 --k 5
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from sklearn.ensemble import HistGradientBoostingClassifier
from sklearn.metrics import roc_auc_score

from weichenanalyse.labels import DEFAULT_META, load_confirmed_faults, load_switches
from weichenanalyse.model import FEATURE_COLS, assemble_features


def healthy_object_ids(switches):
    ws = load_workbook(DEFAULT_META.parent / "labels" / "stoerungen.xlsx", data_only=True)["Gesund_bestaetigt"]
    sw = switches.copy()
    sw["key"] = sw["har_file"].str.replace(".har", "", regex=False).str.strip()
    keys = [str(r[0]).replace(".har", "").strip()
            for r in list(ws.iter_rows(values_only=True))[2:]
            if r[0] and "Dateiname" not in str(r[0])]
    return list(sw[sw.key.isin(keys)].object_id)


def rolling_max(prob, dt, k):
    """Max des gleitenden Mittels (Fenster k) — bildet 'anhaltend hoch' ab."""
    order = np.argsort(dt)
    s = pd.Series(prob[order]).rolling(k, min_periods=1).mean()
    return float(s.max()), order, s.to_numpy()


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--horizon-days", type=int, default=30)
    ap.add_argument("--k", type=int, default=5, help="Glättungsfenster (Persistenz)")
    args = ap.parse_args()

    feats = assemble_features()
    switches = load_switches()
    healthy = set(healthy_object_ids(switches))
    faults = load_confirmed_faults()
    faults["fd"] = pd.to_datetime(faults["datum_beginn"])
    faults = faults[~faults["notiz"].astype(str).str.contains("Stein", na=False)]
    faults = faults[faults.object_id.notna() & faults.fd.notna()].drop_duplicates("object_id")
    fault_fd = dict(zip(faults.object_id, faults.fd))

    labeled = set(fault_fd) | healthy
    df = feats[feats.object_id.isin(labeled)].copy()

    # Labels + Zeitfilter
    H = pd.Timedelta(days=args.horizon_days)
    keep_rows, labels = [], []
    for r in df.itertuples(index=False):
        if r.object_id in fault_fd:
            fd = fault_fd[r.object_id]
            if r.dt > fd:           # Umläufe nach der Reparatur verwerfen
                keep_rows.append(False); labels.append(0); continue
            labels.append(int(r.dt >= fd - H)); keep_rows.append(True)
        else:
            labels.append(0); keep_rows.append(True)
    df["label"] = labels
    df = df[pd.Series(keep_rows, index=df.index)].reset_index(drop=True)

    X = df[FEATURE_COLS].fillna(0.0)
    y = df["label"].to_numpy()

    # LOSO
    rows = []
    sw_ids = sorted(labeled)
    for s in sw_ids:
        tr = df.object_id != s
        ytr = y[tr.to_numpy()]
        if ytr.sum() == 0 or ytr.sum() == len(ytr):
            cw = None
        # Balanciertes Sample-Gewicht
        w = np.where(ytr == 1, (ytr == 0).sum() / max((ytr == 1).sum(), 1), 1.0)
        clf = HistGradientBoostingClassifier(max_depth=3, learning_rate=0.1,
                                             max_iter=200, l2_regularization=1.0)
        clf.fit(X[tr.to_numpy()], ytr, sample_weight=w)

        g = df[df.object_id == s]
        prob = clf.predict_proba(X[(df.object_id == s).to_numpy()])[:, 1]
        is_fault = s in fault_fd
        if is_fault:
            mask = (g.dt <= fault_fd[s]).to_numpy()
        else:
            mask = np.ones(len(g), dtype=bool)
        if mask.sum() == 0:
            continue
        score, order, roll = rolling_max(prob[mask], g.dt.to_numpy()[mask], args.k)
        rows.append({"object_id": s, "is_fault": is_fault, "score": score})

    res = pd.DataFrame(rows)
    yv = res.is_fault.astype(int).to_numpy()
    auc = roc_auc_score(yv, res.score.to_numpy())
    print(f"=== Lernendes Modell, LOSO (horizon={args.horizon_days}d, k={args.k}) ===")
    print(f"Weichen: {int(yv.sum())} Störungen + {int((1-yv).sum())} gesund")
    print(f"Switch-Level ROC-AUC (Störung vs. gesund): {auc:.3f}\n")
    print(f"{'Schwelle':>8} {'Recall':>8} {'Fehlalarm':>10}")
    for thr in np.quantile(res.score, [0.5, 0.6, 0.7, 0.8, 0.9]):
        rec = res[(res.is_fault) & (res.score >= thr)].shape[0] / max(yv.sum(), 1)
        fa = res[(~res.is_fault) & (res.score >= thr)].shape[0] / max((1 - yv).sum(), 1)
        print(f"{thr:8.3f} {rec:8.0%} {fa:10.0%}")
    print("\nBaseline (Handregel, LOSO): Recall 52% / Fehlalarm 61%")


if __name__ == "__main__":
    main()
