"""Ensemble-Vorwarnung gegen bestätigte Störungen + Fehlalarm-Check auf Gesunden.

Mehrere per-Weiche-relative Persistenz-Warner (Strom/Leistung über `mean_amp`,
`peak_amp`, `turn_time`), ODER-verknüpft. Abrupte Fehler (Stein) sind ausgeklammert.

Usage:
    python scripts/eval_warning.py                       # Standard z=2, min_consecutive=10
    python scripts/eval_warning.py --z 1.5 --min-consecutive 5
    python scripts/eval_warning.py --sweep               # Recall/Fehlalarm-Tabelle
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import pandas as pd
from openpyxl import load_workbook

from weichenanalyse.labels import DEFAULT_META, load_confirmed_faults, load_labeled_dataset, load_switches
from weichenanalyse.shape import SHAPE_COLS, attach_shape_features
from weichenanalyse.warning import EnsembleWarning, alarm_active

# Amplitude + Umlaufzeit + Form-Merkmale (Laufphasen-Steigung, Energie, ...).
FEATURES = ("mean_amp", "peak_amp", "turn_time", *SHAPE_COLS)


def healthy_object_ids(switches: pd.DataFrame) -> set:
    xlsx = DEFAULT_META.parent / "labels" / "stoerungen.xlsx"
    ws = load_workbook(xlsx, data_only=True)["Gesund_bestaetigt"]
    sw = switches.copy()
    sw["key"] = sw["har_file"].str.replace(".har", "", regex=False).str.strip()
    keys = [str(r[0]).replace(".har", "").strip()
            for r in list(ws.iter_rows(values_only=True))[2:]
            if r[0] and "Dateiname" not in str(r[0])]
    return set(sw[sw.key.isin(keys)].object_id)


def evaluate(meta, faults, healthy, z, mc, tail=0):
    ens = EnsembleWarning(features=FEATURES, z_thresh=z, min_consecutive=mc).fit(meta)
    pred = ens.predict(meta)
    pred["dt"] = pd.to_datetime(pred["time"], unit="ms")
    hits = tot = 0
    detail = []
    for _, r in faults.iterrows():
        if pd.isna(r.object_id):
            continue
        tot += 1
        g = pred[pred.object_id == r.object_id].sort_values("dt")
        upto = g[g.dt <= r["fd"]]           # Bewertung bis zum Vorfalldatum
        ok = alarm_active(upto, tail=tail)
        first_warn = upto[upto.warn]["dt"].min() if upto.warn.any() else None
        lead = (r["fd"] - first_warn).days if ok and first_warn is not None else None
        hits += int(ok)
        detail.append((r["weiche"], g["signal_unit"].iloc[0] if len(g) else "?", ok, lead))
    fa_healthy = sum(alarm_active(pred[pred.object_id == oid], tail=tail) for oid in healthy)
    return hits, tot, fa_healthy, len(healthy), detail


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--z", type=float, default=1.5)
    ap.add_argument("--min-consecutive", type=int, default=5)
    ap.add_argument("--tail", type=int, default=0,
                    help="Debounce: Alarm nur, wenn in den letzten N Umläufen aktiv (0=aus)")
    ap.add_argument("--sweep", action="store_true")
    args = ap.parse_args()

    meta = load_labeled_dataset(horizon=0)
    meta = attach_shape_features(meta)
    meta["dt"] = pd.to_datetime(meta["time"], unit="ms")
    switches = load_switches()
    healthy = healthy_object_ids(switches)
    faults = load_confirmed_faults()
    faults["fd"] = pd.to_datetime(faults["datum_beginn"])
    # Abrupte Fehler ausklammern (nicht vorwarnbar)
    faults = faults[~faults["notiz"].astype(str).str.contains("Stein", na=False)]

    if args.sweep:
        print(f"{'z':>4} {'min_run':>7} {'tail':>5} {'Recall':>10} {'Fehlalarm':>10}")
        for z, mc, tail in [(2.0, 10, 0), (1.5, 5, 0), (1.5, 5, 40),
                            (1.5, 5, 20), (1.5, 3, 40)]:
            h, t, fh, nh, _ = evaluate(meta, faults, healthy, z, mc, tail)
            print(f"{z:>4} {mc:>7} {tail:>5} {f'{h}/{t}':>10} {f'{fh}/{nh}':>10}")
        return

    h, t, fh, nh, detail = evaluate(meta, faults, healthy, args.z, args.min_consecutive, args.tail)
    print(f"Konfig: z>{args.z}, min_consecutive={args.min_consecutive}, tail={args.tail}, features={FEATURES}")
    print(f"\nRecall (graduelle Störungen): {h}/{t}   |   Fehlalarm gesund: {fh}/{nh}\n")
    print(f"{'Weiche':14s} {'Unit':4s} {'gewarnt':>7s} {'Lead(d)':>7s}")
    for w, unit, ok, lead in detail:
        print(f"{str(w):14s} {unit:4s} {('JA' if ok else 'nein'):>7s} {str(lead) if lead is not None else '—':>7s}")


if __name__ == "__main__":
    main()
