"""1D-CNN auf der Rohkurve, Leave-one-switch-out — direkt vergleichbar mit dem GBM.

Eingang: per-Weiche normierte, resampelte Wellenform (Strom oder Leistung).
Label/Eval identisch zu scripts/loso_model.py (Fenster vor Vorfall, gesunde negativ,
geglätteter Score, Switch-Level ROC-AUC + Recall/Fehlalarm). So ist 'klassisch vs Deep'
auf identischer Basis bewertbar.

Usage:
    python scripts/loso_cnn.py --horizon-days 60 --k 5 --predictable-only
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import numpy as np
import pandas as pd
import torch
import torch.nn as nn
from openpyxl import load_workbook
from sklearn.metrics import roc_auc_score

from weichenanalyse.labels import (DEFAULT_META, PREDICTABLE_TYPES, categorize_fault,
                                   load_confirmed_faults, load_switches)
from weichenanalyse.waveforms import waveform_matrix


class CNN(nn.Module):
    def __init__(self):
        super().__init__()
        self.net = nn.Sequential(
            nn.Conv1d(1, 16, 7, padding=3), nn.ReLU(), nn.MaxPool1d(2),
            nn.Conv1d(16, 32, 5, padding=2), nn.ReLU(), nn.AdaptiveAvgPool1d(1),
        )
        self.fc = nn.Linear(32, 1)

    def forward(self, x):
        return self.fc(self.net(x).squeeze(-1)).squeeze(-1)


def healthy_object_ids(switches):
    ws = load_workbook(DEFAULT_META.parent / "labels" / "stoerungen.xlsx", data_only=True)["Gesund_bestaetigt"]
    sw = switches.copy()
    sw["key"] = sw["har_file"].str.replace(".har", "", regex=False).str.strip()
    keys = [str(r[0]).replace(".har", "").strip()
            for r in list(ws.iter_rows(values_only=True))[2:]
            if r[0] and "Dateiname" not in str(r[0])]
    return list(sw[sw.key.isin(keys)].object_id)


def train_predict(Xtr, ytr, Xte, epochs=8, seed=0):
    torch.manual_seed(seed)
    dev = "cpu"
    Xt = torch.tensor(Xtr).unsqueeze(1)
    yt = torch.tensor(ytr.astype("float32"))
    pos_w = torch.tensor([(ytr == 0).sum() / max((ytr == 1).sum(), 1)], dtype=torch.float32)
    model = CNN().to(dev)
    opt = torch.optim.Adam(model.parameters(), lr=1e-3)
    loss_fn = nn.BCEWithLogitsLoss(pos_weight=pos_w)
    n = len(Xt)
    for _ in range(epochs):
        perm = torch.randperm(n)
        for i in range(0, n, 512):
            idx = perm[i:i + 512]
            opt.zero_grad()
            loss = loss_fn(model(Xt[idx]), yt[idx])
            loss.backward()
            opt.step()
    model.eval()
    with torch.no_grad():
        p = torch.sigmoid(model(torch.tensor(Xte).unsqueeze(1))).numpy()
    return p


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--horizon-days", type=int, default=60)
    ap.add_argument("--k", type=int, default=5)
    ap.add_argument("--epochs", type=int, default=8)
    ap.add_argument("--predictable-only", action="store_true")
    args = ap.parse_args()

    keys, X = waveform_matrix()
    switches = load_switches()
    healthy = set(healthy_object_ids(switches))
    faults = load_confirmed_faults()
    faults["fd"] = pd.to_datetime(faults["datum_beginn"])
    faults["cat"] = faults["notiz"].apply(categorize_fault)
    faults = faults[faults.cat != "Stein"]
    if "ursache" in faults.columns:
        faults = faults[faults["ursache"].astype(str).str.lower() != "inspektion"]
    if args.predictable_only:
        faults = faults[faults.cat.isin(PREDICTABLE_TYPES)]
    faults = faults[faults.object_id.notna() & faults.fd.notna()].drop_duplicates("object_id")
    fault_fd = dict(zip(faults.object_id, faults.fd))
    fault_cat = dict(zip(faults.object_id, faults.cat))

    labeled = set(fault_fd) | healthy
    mask = keys.object_id.isin(labeled).to_numpy()
    keys, X = keys[mask].reset_index(drop=True), X[mask]

    H = pd.Timedelta(days=args.horizon_days)
    y = np.zeros(len(keys), dtype=int)
    drop = np.zeros(len(keys), dtype=bool)
    for i, r in enumerate(keys.itertuples(index=False)):
        if r.object_id in fault_fd:
            fd = fault_fd[r.object_id]
            if r.dt > fd:
                drop[i] = True
            elif r.dt >= fd - H:
                y[i] = 1
    keep = ~drop
    keys, X, y = keys[keep].reset_index(drop=True), X[keep], y[keep]

    rows = []
    for s in sorted(labeled):
        tr = (keys.object_id != s).to_numpy()
        te = (keys.object_id == s).to_numpy()
        if y[tr].sum() == 0:
            continue
        prob = train_predict(X[tr], y[tr], X[te], epochs=args.epochs)
        g = keys[te]
        is_fault = s in fault_fd
        m = (g.dt <= fault_fd[s]).to_numpy() if is_fault else np.ones(len(g), dtype=bool)
        if m.sum() == 0:
            continue
        order = np.argsort(g.dt.to_numpy()[m])
        roll = pd.Series(prob[m][order]).rolling(args.k, min_periods=1).mean()
        rows.append({"object_id": s, "is_fault": is_fault, "score": float(roll.max()),
                     "cat": fault_cat.get(s, "gesund")})

    res = pd.DataFrame(rows)
    yv = res.is_fault.astype(int).to_numpy()
    auc = roc_auc_score(yv, res.score.to_numpy())
    scope = "vorhersagbare Familie" if args.predictable_only else "alle graduellen"
    print(f"=== 1D-CNN, LOSO ({scope}; horizon={args.horizon_days}d, k={args.k}, epochs={args.epochs}) ===")
    print(f"Weichen: {int(yv.sum())} Störungen + {int((1-yv).sum())} gesund")
    print(f"Switch-Level ROC-AUC: {auc:.3f}")
    thr = np.quantile(res[~res.is_fault].score, 0.80)
    rec = res[(res.is_fault) & (res.score >= thr)].shape[0] / max(yv.sum(), 1)
    fa = res[(~res.is_fault) & (res.score >= thr)].shape[0] / max((1 - yv).sum(), 1)
    print(f"Betriebspunkt (~20% Fehlalarm): Recall {rec:.0%}, Fehlalarm {fa:.0%}\n")
    print("Recall je Fehlertyp:")
    for cat, gg in res[res.is_fault].groupby("cat"):
        print(f"  {cat:12s} {(gg.score >= thr).sum()}/{len(gg)}")
    print("\nVergleich GBM (bereinigt, gleiche Basis): AUC 0.79 / Recall 67% @20% FA")


if __name__ == "__main__":
    main()
